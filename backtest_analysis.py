import pandas as pd
from openpyxl import load_workbook

# ---------- File paths ----------
backtest_file = "backtest_df.xlsx"
features_file = "feature_backtest.csv"
output_file   = "backtest_df_updated.xlsx"

# ---------- 1. Read the two sources ----------
df_backtest = pd.read_excel(backtest_file, sheet_name="Sheet1", dtype=str)
df_features = pd.read_csv(features_file, dtype=str)

# ---------- 2. Find and rename key columns ----------
# Adjust the fallback indices only if you know the layout is different
def find_col_by_names(df, preferred, fallback_idx=None):
    for nm in preferred:
        for c in df.columns:
            if nm.lower() in str(c).lower():
                return c
    if fallback_idx is not None and fallback_idx < len(df.columns):
        return df.columns[fallback_idx]
    raise KeyError(f"Could not find any of {preferred}")

back_date_col   = find_col_by_names(df_backtest, ['date'], fallback_idx=1)
back_symbol_col = find_col_by_names(df_backtest, ['symbol','ticker'], fallback_idx=2)
feat_date_col   = find_col_by_names(df_features, ['date'], fallback_idx=7)
feat_symbol_col = find_col_by_names(df_features, ['symbol','ticker'], fallback_idx=6)

df_backtest = df_backtest.rename(columns={back_date_col:'date', back_symbol_col:'symbol'})
df_features = df_features.rename(columns={feat_date_col:'date', feat_symbol_col:'symbol'})

# ---------- 3. Clean key columns ----------
def clean_symbol(s):
    return (s.fillna('')
             .astype(str)
             .str.replace('\xa0',' ',regex=False)
             .str.replace('\u200b','',regex=False)
             .str.replace(r'\s+',' ',regex=True)
             .str.strip()
             .str.upper())

df_backtest['symbol'] = clean_symbol(df_backtest['symbol'])
df_features['symbol'] = clean_symbol(df_features['symbol'])

def parse_dates_auto(s):
    a = pd.to_datetime(s, errors='coerce', dayfirst=False)
    if a.isna().mean() > 0.05:
        b = pd.to_datetime(s, errors='coerce', dayfirst=True)
        return a if a.isna().sum() <= b.isna().sum() else b
    return a

df_backtest['date'] = parse_dates_auto(df_backtest['date']).dt.normalize()
df_features['date'] = parse_dates_auto(df_features['date']).dt.normalize()

# ---------- 4. Build a merge key to avoid dtype problems ----------
df_backtest['merge_key'] = df_backtest['symbol'] + '|' + df_backtest['date'].dt.strftime('%Y-%m-%d')
df_features['merge_key'] = df_features['symbol'] + '|' + df_features['date'].dt.strftime('%Y-%m-%d')

# ---------- 5. Select only real feature columns ----------
feature_cols = [c for c in df_features.columns
                if c not in ('date','symbol','merge_key')
                and not c.lower().startswith('unnamed')]

df_feat_sample = df_features[['merge_key'] + feature_cols]

# ---------- 6. Merge ----------
print(f"Rows in backtest : {len(df_backtest)}")
tmp = df_backtest.merge(df_feat_sample[['merge_key']], on='merge_key',
                        how='left', indicator=True)
print("Rows with NO feature match :",
      (tmp['_merge']=='left_only').sum())

df_merged = df_backtest.merge(df_feat_sample, on='merge_key', how='left')
df_merged = df_merged.drop(columns=['merge_key'])

# ---------- 7. Place feature columns from column N (index 13) ----------
N_index = 13
left_part  = df_merged.iloc[:, :N_index]
right_part = df_merged[feature_cols]
df_final   = pd.concat([left_part.reset_index(drop=True),
                        right_part.reset_index(drop=True)], axis=1)

# ---------- 8. (Optional) Check how many true NaN values are present ----------
print("\nMissing values per feature column (NaN counts):")
print(df_final[feature_cols].isna().sum().sort_values(ascending=False).head(20))

# ---------- 9. Save to Excel ----------
df_final.to_excel(output_file, index=False)

# ---------- 10. Post-processing in Excel ----------
wb = load_workbook(output_file)
ws = wb.active

col_action  = 4   # D
col_symbol  = 3   # C
col_pnl_pct = 12  # L

max_row = ws.max_row
rows_to_delete = []

for row in range(2, max_row + 1):
    action_value = ws.cell(row=row, column=col_action).value
    if action_value == "BUY":
        symbol = ws.cell(row=row, column=col_symbol).value
        found_value = None
        for search_row in range(row + 1, min(max_row, row + 1500) + 1):
            if ws.cell(search_row, column=col_symbol).value == symbol:
                found_value = ws.cell(search_row, column=col_pnl_pct).value
                break
        if found_value is not None:
            ws.cell(row=row, column=col_pnl_pct).value = found_value
    elif action_value == "SELL":
        rows_to_delete.append(row)

for r in reversed(rows_to_delete):
    ws.delete_rows(r, 1)

ws.insert_rows(2, amount=2)

# ---- summary rows: min/max/avg of feature columns grouped by pnl>0 or <0 ----
import numpy as np
data = pd.DataFrame(ws.values)
data.columns = data.iloc[0]
data = data.drop(0)
data = data.apply(pd.to_numeric, errors="ignore")

pnl = pd.to_numeric(data.iloc[:, col_pnl_pct-1], errors="coerce")
mask_pos, mask_neg = pnl > 0, pnl < 0
max_col = ws.max_column

for col in range(N_index, max_col + 1):
    col_values = pd.to_numeric(data.iloc[:, col-1], errors="coerce")
    vals_pos = col_values[mask_pos].dropna()
    if not vals_pos.empty:
        ws.cell(row=2, column=col).value = \
            f"{vals_pos.min():.2f}, {vals_pos.max():.2f}, {vals_pos.mean():.2f}"
    vals_neg = col_values[mask_neg].dropna()
    if not vals_neg.empty:
        ws.cell(row=3, column=col).value = \
            f"{vals_neg.min():.2f}, {vals_neg.max():.2f}, {vals_neg.mean():.2f}"

wb.save(output_file)
print(f"\n✅ Final workbook saved to {output_file}")
